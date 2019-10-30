# -*- coding: utf-8 -*-
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import xlwt
import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles import colors
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.drawing.image import Image

input_file_name = input('Please provide the Excel file name(strict format: name.xlsx):')

dd=pd.read_excel(open(input_file_name,'rb'),sheet_name='Master CP List')
Ticker= dd.iloc[5:(len(dd)),1]
Issuer=dd.iloc[5:(len(dd)),0]
DTM =dd.iloc[5:len(dd),2]
Maturity=DTM
Maturity_date =dd.iloc[5:len(dd),3]
sp=dd.iloc[5:len(dd),4]
M=dd.iloc[5:len(dd),5]
Disc= dd.iloc[5:len(dd),8]

data2 =pd.concat([Ticker,DTM,Maturity,Disc,sp,Issuer,Maturity_date,M],axis=1)
data2.columns =['Ticker','DTM','Maturity','Disc','S&P','Issuer name','M Date','M']

#new=data2['DTM'].str.split(' - ',n=1,expand=True)


#choose the first number of the DTM
for i in range(len(data2)):
          if type(data2.iloc[i,1]) is str:

               data2.iloc[i,1]=(data2.iloc[i,1])[0:3]
          else:
              data2.iloc[i, 1]=data2.iloc[i, 1]


for i in range(len(data2)):
       if type(data2.iloc[i,1]) is str:
           if   '-' not in data2.iloc[i,1]:

               data2.iloc[i,1]=data2.iloc[i, 1]
           else:
              data2.iloc[i, 1]=(data2.iloc[i, 1])[0:1]

for i in range(len(data2)):
    if type(data2.iloc[i,1]) is str:
        data2.iloc[i, 1] = int(data2.iloc[i, 1])
    else:
        data2.iloc[i, 1] = data2.iloc[i, 1]


#choose the first date of the Maturity date
for i in range(len(data2)):
    if '-' not in data2.iloc[i,6]:
        data2.iloc[i, 6] = data2.iloc[i, 6]
    else:
        data2.iloc[i, 6] = (data2.iloc[i, 6])[0:5]




# read file and count the amount of bonds in every bond level
df =data2



#Get maturity in format of days in period#
for i in range(0,len(df),1):


   if df.iloc[i,1]==1 :
     df.iloc[i, 2] ='Overnight'

   elif 2<=df.iloc[i,1]<=30 :
    df.iloc[i, 2] ='2-30 days'
   elif 30<df.iloc[i,1]<=60:
    df.iloc[i, 2] = '31-60 days'
   elif 60<df.iloc[i,1]<=90:
          df.iloc[i, 2] ='61-90 days'
   elif 90 < df.iloc[i, 1] <= 120:
     df.iloc[i, 2] = '91-120 days'
   elif 120 < df.iloc[i, 1] <= 180:
     df.iloc[i, 2] = '121-180 days'
   elif 180 < df.iloc[i, 1] <= 270:
       df.iloc[i, 2] = '181-270 days'
   else:
        df.iloc[i, 2] ='More than 270 days'


# seperate maturity
m_overnight = df[df['Maturity']=='Overnight']
m_2_30 =  df[df['Maturity']=='2-30 days']
m_31_60 =  df[df['Maturity']=='31-60 days']
m_61_90 =  df[df['Maturity']=='61-90 days']
m_91_120 =  df[df['Maturity']=='91-120 days']
m_121_180 =  df[df['Maturity']=='121-180 days']
m_181_270 =  df[df['Maturity']=='181-270 days']
m_270_more =  df[df['Maturity']=='More than 270 days']

# get SP count

def get_count(bond):
    return pd.value_counts(bond)



#print the  sp count
print('overnight')
print(get_count(m_overnight['S&P']))
print('2-30 days')
print(get_count(m_2_30['S&P']))
print('31-60 days')
print(get_count(m_31_60['S&P']))
print('61-90 days')
print(get_count(m_61_90['S&P']))
print('91-120 days')
print(get_count(m_91_120['S&P']))
print('121-180 days')
print(get_count(m_121_180['S&P']))
print('181-270 days')
print(get_count(m_181_270['S&P']))
print('More than 270 days')
print(get_count(m_270_more['S&P']))



### overnight
m_overnight_a1_plus= m_overnight[m_overnight['S&P']=='A-1+']
m_overnight_a1 =m_overnight[m_overnight['S&P']=='A-1']
m_overnight_a2 =m_overnight[m_overnight['S&P']=='A-2']
m_overnight_a3 =m_overnight[m_overnight['S&P']=='A-3']




###  2～30
m_2_30_a1_plus= m_2_30[m_2_30['S&P']=='A-1+']
m_2_30_a1 =m_2_30[m_2_30['S&P']=='A-1']
m_2_30_a2 =m_2_30[m_2_30['S&P']=='A-2']
m_2_30_a3 =m_2_30[m_2_30['S&P']=='A-3']


# 31～60
m_31_60_a1_plus=m_31_60[m_31_60['S&P']=='A-1+']
m_31_60_a1=m_31_60[m_31_60['S&P']=='A-1']
m_31_60_a2= m_31_60[m_31_60['S&P']=='A-2']
m_31_60_a3=m_31_60[m_31_60['S&P']=='A-3']

#61～90
m_61_90_a1_plus=m_61_90[m_61_90['S&P']=='A-1+']
m_61_90_a1=m_61_90[m_61_90['S&P']=='A-1']
m_61_90_a2= m_61_90[m_61_90['S&P']=='A-2']
m_61_90_a3=m_61_90[m_61_90['S&P']=='A-3']


#91～120

m_91_120_a1_plus=m_91_120[m_91_120['S&P']=='A-1+']
m_91_120_a1=m_91_120[m_91_120['S&P']=='A-1']
m_91_120_a2= m_91_120[m_91_120['S&P']=='A-2']
m_91_120_a3=m_91_120[m_91_120['S&P']=='A-3']
#121～180

m_121_180_a1_plus=m_121_180[m_121_180['S&P']=='A-1+']
m_121_180_a1=m_121_180[m_121_180['S&P']=='A-1']
m_121_180_a2= m_121_180[m_121_180['S&P']=='A-2']
m_121_180_a3=m_121_180[m_121_180['S&P']=='A-3']


#181～270   _
m_181_270_a1_plus=m_181_270[m_181_270['S&P']=='A-1+']
m_181_270_a1=m_181_270[m_181_270['S&P']=='A-1']
m_181_270_a2= m_181_270[m_181_270['S&P']=='A-2']
m_181_270_a3=m_181_270[m_181_270['S&P']=='A-3']



#270 more

m_270_more_a1_plus=m_270_more[m_270_more['S&P']=='A-1+']
m_270_more_a1=m_270_more[m_270_more['S&P']=='A-1']
m_270_more_a2= m_270_more[m_270_more['S&P']=='A-2']
m_270_more_a3=m_270_more[m_270_more['S&P']=='A-3']

# sort by disc
def sort_disc(bond):
    bond=bond.sort_values(by=['Disc'],ascending=False)
    return bond

def drop_dup(bond):
    bond = bond.drop_duplicates('Ticker')
    return bond

def get_top10(bond):
       bond = bond.iloc[:10]
       return bond

def drop_T_and_D(bond):
    bond = bond.drop(['Ticker'],axis=1)
    return bond
def reorder_column(bond):
        bond= bond.loc[:,['S&P','M','Issuer name','DTM', 'M Date','Disc', 'Maturity']]
        return bond

def pivot_bond(bond):
    bond = pd.pivot_table(bond, values='Disc', index=['S&P', 'Issuer name', 'DTM','Maturity date'], columns='Maturity')
    return bond

def sort_disc_again(bond):
    bond = bond.sort_values(by=[list(bond)[0]], ascending=False)
    return bond


# 0.(m_overnight) sort, drop_duplicates, get_top10(overnight), drop "ticker",re_order
m_overnight_a1_plus=sort_disc(m_overnight_a1_plus)
m_overnight_a1 =sort_disc(m_overnight_a1)
m_overnight_a2 =sort_disc(m_overnight_a2)
m_overnight_a3 =sort_disc(m_overnight_a3)


m_overnight_a1_plus=drop_dup(m_overnight_a1_plus)
m_overnight_a1 =drop_dup(m_overnight_a1)
m_overnight_a2 =drop_dup(m_overnight_a2)
m_overnight_a3 =drop_dup(m_overnight_a3)


m_overnight_a1_plus=get_top10(m_overnight_a1_plus)
m_overnight_a1 =get_top10(m_overnight_a1)
m_overnight_a2 =get_top10(m_overnight_a2)
m_overnight_a3 =get_top10(m_overnight_a3)

m_overnight_a1_plus=drop_T_and_D(m_overnight_a1_plus)
m_overnight_a1 =drop_T_and_D(m_overnight_a1)
m_overnight_a2 =drop_T_and_D(m_overnight_a2)
m_overnight_a3 =drop_T_and_D(m_overnight_a3)


m_overnight_a1_plus=reorder_column(m_overnight_a1_plus)
m_overnight_a1=reorder_column(m_overnight_a1)
m_overnight_a2=reorder_column(m_overnight_a2)
m_overnight_a3=reorder_column(m_overnight_a3)




#1. (m_0_30)sort, drop_duplicates, get_top10(overnight), drop "ticker",'DTM',re_order
m_2_30_a1_plus=sort_disc(m_2_30_a1_plus)
m_2_30_a1 =sort_disc(m_2_30_a1)
m_2_30_a2 =sort_disc(m_2_30_a2)
m_2_30_a3 =sort_disc(m_2_30_a3)



m_2_30_a1_plus=drop_dup(m_2_30_a1_plus)
m_2_30_a1 =drop_dup(m_2_30_a1)
m_2_30_a2 =drop_dup(m_2_30_a2)
m_2_30_a3 =drop_dup(m_2_30_a3)


m_2_30_a1_plus=get_top10(m_2_30_a1_plus)
m_2_30_a1 =get_top10(m_2_30_a1)
m_2_30_a2 =get_top10(m_2_30_a2)
m_2_30_a3 =get_top10(m_2_30_a3)


m_2_30_a1_plus=drop_T_and_D(m_2_30_a1_plus)
m_2_30_a1 =drop_T_and_D(m_2_30_a1)
m_2_30_a2 =drop_T_and_D(m_2_30_a2)
m_2_30_a3 =drop_T_and_D(m_2_30_a3)



m_2_30_a1_plus=reorder_column(m_2_30_a1_plus)
m_2_30_a1=reorder_column(m_2_30_a1)
m_2_30_a2=reorder_column(m_2_30_a2)
m_2_30_a3=reorder_column(m_2_30_a3)



#2.(m_31_60) sort, drop_duplicates, get_top10(overnight), drop "ticker",'DTM',re_order
m_31_60_a1_plus=sort_disc(m_31_60_a1_plus)
m_31_60_a1 =sort_disc(m_31_60_a1)
m_31_60_a2 =sort_disc(m_31_60_a2)
m_31_60_a3 =sort_disc(m_31_60_a3)



m_31_60_a1_plus=drop_dup(m_31_60_a1_plus)
m_31_60_a1 =drop_dup(m_31_60_a1)
m_31_60_a2 =drop_dup(m_31_60_a2)
m_31_60_a3 =drop_dup(m_31_60_a3)


m_31_60_a1_plus=get_top10(m_31_60_a1_plus)
m_31_60_a1 =get_top10(m_31_60_a1)
m_31_60_a2 =get_top10(m_31_60_a2)
m_31_60_a3 =get_top10(m_31_60_a3)


m_31_60_a1_plus=drop_T_and_D(m_31_60_a1_plus)
m_31_60_a1 =drop_T_and_D(m_31_60_a1)
m_31_60_a2 =drop_T_and_D(m_31_60_a2)
m_31_60_a3 =drop_T_and_D(m_31_60_a3)


m_31_60_a1_plus=reorder_column(m_31_60_a1_plus)
m_31_60_a1=reorder_column(m_31_60_a1)
m_31_60_a2=reorder_column(m_31_60_a2)
m_31_60_a3=reorder_column(m_31_60_a3)


#3.(m_61_90)sort, drop_duplicates, get_top10(overnight), drop "ticker",'DTM',re_order
m_61_90_a1_plus=sort_disc(m_61_90_a1_plus)
m_61_90_a1 =sort_disc(m_61_90_a1)
m_61_90_a2 =sort_disc(m_61_90_a2)
m_61_90_a3 =sort_disc(m_61_90_a3)



m_61_90_a1_plus=drop_dup(m_61_90_a1_plus)
m_61_90_a1 =drop_dup(m_61_90_a1)
m_61_90_a2 =drop_dup(m_61_90_a2)
m_61_90_a3 =drop_dup(m_61_90_a3)


m_61_90_a1_plus=get_top10(m_61_90_a1_plus)
m_61_90_a1 =get_top10(m_61_90_a1)
m_61_90_a2 =get_top10(m_61_90_a2)
m_61_90_a3 =get_top10(m_61_90_a3)


m_61_90_a1_plus=drop_T_and_D(m_61_90_a1_plus)
m_61_90_a1 =drop_T_and_D(m_61_90_a1)
m_61_90_a2 =drop_T_and_D(m_61_90_a2)
m_61_90_a3 =drop_T_and_D(m_61_90_a3)


m_61_90_a1_plus=reorder_column(m_61_90_a1_plus)
m_61_90_a1=reorder_column(m_61_90_a1)
m_61_90_a2=reorder_column(m_61_90_a2)
m_61_90_a3=reorder_column(m_61_90_a3)

#4.(m_91_120)sort, drop_duplicates, get_top10(overnight), drop "ticker",'DTM',re_order
m_91_120_a1_plus=sort_disc(m_91_120_a1_plus)
m_91_120_a1 =sort_disc(m_91_120_a1)
m_91_120_a2 =sort_disc(m_91_120_a2)
m_91_120_a3 =sort_disc(m_91_120_a3)



m_91_120_a1_plus=drop_dup(m_91_120_a1_plus)
m_91_120_a1 =drop_dup(m_91_120_a1)
m_91_120_a2 =drop_dup(m_91_120_a2)
m_91_120_a3 =drop_dup(m_91_120_a3)


m_91_120_a1_plus=get_top10(m_91_120_a1_plus)
m_91_120_a1 =get_top10(m_91_120_a1)
m_91_120_a2 =get_top10(m_91_120_a2)
m_91_120_a3 =get_top10(m_91_120_a3)


m_91_120_a1_plus=drop_T_and_D(m_91_120_a1_plus)
m_91_120_a1 =drop_T_and_D(m_91_120_a1)
m_91_120_a2 =drop_T_and_D(m_91_120_a2)
m_91_120_a3 =drop_T_and_D(m_91_120_a3)


m_91_120_a1_plus=reorder_column(m_91_120_a1_plus)
m_91_120_a1=reorder_column(m_91_120_a1)
m_91_120_a2=reorder_column(m_91_120_a2)
m_91_120_a3=reorder_column(m_91_120_a3)

#5.(m_121_180)sort, drop_duplicates, get_top10(overnight), drop "ticker",'DTM',re_order
m_121_180_a1_plus=sort_disc(m_121_180_a1_plus)
m_121_180_a1 =sort_disc(m_121_180_a1)
m_121_180_a2 =sort_disc(m_121_180_a2)
m_121_180_a3 =sort_disc(m_121_180_a3)



m_121_180_a1_plus=drop_dup(m_121_180_a1_plus)
m_121_180_a1 =drop_dup(m_121_180_a1)
m_121_180_a2 =drop_dup(m_121_180_a2)
m_121_180_a3 =drop_dup(m_121_180_a3)


m_121_180_a1_plus=get_top10(m_121_180_a1_plus)
m_121_180_a1 =get_top10(m_121_180_a1)
m_121_180_a2 =get_top10(m_121_180_a2)
m_121_180_a3 =get_top10(m_121_180_a3)


m_121_180_a1_plus=drop_T_and_D(m_121_180_a1_plus)
m_121_180_a1 =drop_T_and_D(m_121_180_a1)
m_121_180_a2 =drop_T_and_D(m_121_180_a2)
m_121_180_a3 =drop_T_and_D(m_121_180_a3)


m_121_180_a1_plus=reorder_column(m_121_180_a1_plus)
m_121_180_a1=reorder_column(m_121_180_a1)
m_121_180_a2=reorder_column(m_121_180_a2)
m_121_180_a3=reorder_column(m_121_180_a3)

#6.(m_181_270)sort, drop_duplicates, get_top10(overnight), drop "ticker",'DTM',re_order
m_181_270_a1_plus=sort_disc(m_181_270_a1_plus)
m_181_270_a1 =sort_disc(m_181_270_a1)
m_181_270_a2 =sort_disc(m_181_270_a2)
m_181_270_a3 =sort_disc(m_181_270_a3)



m_181_270_a1_plus=drop_dup(m_181_270_a1_plus)
m_181_270_a1 =drop_dup(m_181_270_a1)
m_181_270_a2 =drop_dup(m_181_270_a2)
m_181_270_a3 =drop_dup(m_181_270_a3)


m_181_270_a1_plus=get_top10(m_181_270_a1_plus)
m_181_270_a1 =get_top10(m_181_270_a1)
m_181_270_a2 =get_top10(m_181_270_a2)
m_181_270_a3 =get_top10(m_181_270_a3)


m_181_270_a1_plus=drop_T_and_D(m_181_270_a1_plus)
m_181_270_a1 =drop_T_and_D(m_181_270_a1)
m_181_270_a2 =drop_T_and_D(m_181_270_a2)
m_181_270_a3 =drop_T_and_D(m_181_270_a3)


m_181_270_a1_plus=reorder_column(m_181_270_a1_plus)
m_181_270_a1=reorder_column(m_181_270_a1)
m_181_270_a2=reorder_column(m_181_270_a2)
m_181_270_a3=reorder_column(m_181_270_a3)



#4.(m_270_more) sort, drop_duplicates, get_top10(overnight), drop "ticker",'DTM',re_order
m_270_more_a1_plus=sort_disc(m_270_more_a1_plus)
m_270_more_a1 =sort_disc(m_270_more_a1)
m_270_more_a2 =sort_disc(m_270_more_a2)
m_270_more_a3 =sort_disc(m_270_more_a3)



m_270_more_a1_plus=drop_dup(m_270_more_a1_plus)
m_270_more_a1 =drop_dup(m_270_more_a1)
m_270_more_a2 =drop_dup(m_270_more_a2)
m_270_more_a3 =drop_dup(m_270_more_a3)


m_270_more_a1_plus=get_top10(m_270_more_a1_plus)
m_270_more_a1 =get_top10(m_270_more_a1)
m_270_more_a2 =get_top10(m_270_more_a2)
m_270_more_a3 =get_top10(m_270_more_a3)


m_270_more_a1_plus=drop_T_and_D(m_270_more_a1_plus)
m_270_more_a1 =drop_T_and_D(m_270_more_a1)
m_270_more_a2 =drop_T_and_D(m_270_more_a2)
m_270_more_a3 =drop_T_and_D(m_270_more_a3)



m_270_more_a1_plus=reorder_column(m_270_more_a1_plus)
m_270_more_a1=reorder_column(m_270_more_a1)
m_270_more_a2=reorder_column(m_270_more_a2)
m_270_more_a3=reorder_column(m_270_more_a3)

###list of different maturity and different S&P
l1=['m_overnight_a1_plus','m_overnight_a1','m_overnight_a2','m_overnight_a3']
l2=['m_2_30_a1_plus','m_2_30_a1','m_2_30_a2','m_2_30_a3']
l3=['m_31_60_a1_plus','m_31_60_a1','m_31_60_a2' ,'m_31_60_a3']
l4=['m_61_90_a1_plus','m_61_90_a1','m_61_90_a2','m_61_90_a3']
l5=['m_91_120_a1_plus','m_91_120_a1','m_91_120_a2','m_91_120_a3']
l6=['m_121_180_a1_plus','m_121_180_a1','m_121_180_a2','m_121_180_a3']
l7=['m_181_270_a1_plus','m_181_270_a1','m_181_270_a2','m_181_270_a3']
l8=['m_270_more_a1_plus','m_270_more_a1','m_270_more_a2','m_270_more_a3']

#stuff =np.zeros(len(m_overnight_a1_plus.columns),dtype=int)
#$stuff= stuff.fill(np.nan)
stuff =pd.Series([np.nan]*len(m_overnight_a1_plus.columns),index=m_overnight_a1_plus.columns)
## overnight slice
for i in range(12):
    if len(m_overnight_a1_plus)<=10:
        m_overnight_a1_plus =m_overnight_a1_plus.append(stuff,ignore_index=True)
for i in range(12):
    if len(m_overnight_a1)<=10:
        m_overnight_a1 =m_overnight_a1.append(stuff,ignore_index=True)
for i in range(12):
    if len(m_overnight_a2)<=10:
        m_overnight_a2 =m_overnight_a2.append(stuff,ignore_index=True)
for i in range(12):
    if len(m_overnight_a3)<=10:
        m_overnight_a3 =m_overnight_a3.append(stuff,ignore_index=True)
## 2~30 slice
for i in range(12):
    if len(m_2_30_a1_plus)<=10:
        m_2_30_a1_plus =m_2_30_a1_plus.append(stuff,ignore_index=True)

for i in range(12):
    if len(m_2_30_a1)<=10:
        m_2_30_a1 =m_2_30_a1.append(stuff,ignore_index=True)

for i in range(12):
    if len(m_2_30_a2)<=10:
        m_2_30_a2 =m_2_30_a2.append(stuff,ignore_index=True)

for i in range(12):
    if len(m_2_30_a3)<=10:
        m_2_30_a3 =m_2_30_a3.append(stuff,ignore_index=True)
## 31~60 slice
for i in range(12):
    if len(m_31_60_a1_plus)<=10:
        m_31_60_a1_plus =m_31_60_a1_plus.append(stuff,ignore_index=True)

for i in range(12):
    if len(m_31_60_a1)<=10:
        m_31_60_a1 =m_31_60_a1.append(stuff,ignore_index=True)
for i in range(12):
    if len(m_31_60_a2)<=10:
        m_31_60_a2 =m_31_60_a2.append(stuff,ignore_index=True)

for i in range(12):
    if len(m_31_60_a3)<=10:
        m_31_60_a3 =m_31_60_a3.append(stuff,ignore_index=True)

### 61~90 slice
for i in range(12):
    if len(m_61_90_a1_plus)<=10:
        m_61_90_a1_plus = m_61_90_a1_plus.append(stuff, ignore_index=True)
for i in range(12):
    if len(m_61_90_a1) <=10:
        m_61_90_a1 = m_61_90_a1.append(stuff, ignore_index=True)
for i in range(12):
    if len(m_61_90_a2)<=10:
        m_61_90_a2 = m_61_90_a2.append(stuff, ignore_index=True)
for i in range(12):
    if len(m_61_90_a3) <=10:
        m_61_90_a3 = m_61_90_a3.append(stuff, ignore_index=True)
### 91~120 slice
for i in range(12):
    if len(m_91_120_a1_plus)<=10:
        m_91_120_a1_plus = m_91_120_a1_plus.append(stuff, ignore_index=True)
for i in range(12):
    if len(m_91_120_a1) <=10:
        m_91_120_a1 = m_91_120_a1.append(stuff, ignore_index=True)
for i in range(12):
    if len(m_91_120_a2)<=10:
        m_91_120_a2 = m_91_120_a2.append(stuff, ignore_index=True)
for i in range(12):
    if len(m_91_120_a3) <=10:
        m_91_120_a3 = m_91_120_a3.append(stuff, ignore_index=True)

### 121~180 slice
for i in range(12):
    if len(m_121_180_a1_plus)<=10:
        m_121_180_a1_plus = m_121_180_a1_plus.append(stuff, ignore_index=True)
for i in range(12):
    if len(m_121_180_a1) <=10:
        m_121_180_a1 = m_121_180_a1.append(stuff, ignore_index=True)
for i in range(12):
    if len(m_121_180_a2)<=10:
        m_121_180_a2 = m_121_180_a2.append(stuff, ignore_index=True)
for i in range(12):
    if len(m_121_180_a3) <=10:
        m_121_180_a3 = m_121_180_a3.append(stuff, ignore_index=True)
## 181~270 slice

for i in range(12):
    if len(m_181_270_a1_plus)<=10:
        m_181_270_a1_plus = m_181_270_a1_plus.append(stuff, ignore_index=True)

for i in range(12):
    if len(m_181_270_a1) <=10:
        m_181_270_a1 = m_181_270_a1.append(stuff, ignore_index=True)

for i in range(12):
    if len(m_181_270_a2) <=10:
        m_181_270_a2 = m_181_270_a2.append(stuff, ignore_index=True)

for i in range(12):
    if len(m_181_270_a3) <=10:
        m_181_270_a3 = m_181_270_a3.append(stuff, ignore_index=True)
##m 270_more
for i in range(12):
    if len(m_270_more_a1_plus) <=10:
        m_270_more_a1_plus = m_270_more_a1_plus.append(stuff, ignore_index=True)

for i in range(12):
    if len(m_270_more_a1)<=10:
        m_270_more_a1 = m_270_more_a1.append(stuff, ignore_index=True)
for i in range(12):
    if len(m_270_more_a2) <=10:
        m_270_more_a2 = m_270_more_a2.append(stuff, ignore_index=True)

for i in range(12):
    if len(m_270_more_a3)<=10:
        m_270_more_a3 = m_270_more_a3.append(stuff, ignore_index=True)

##concat dataframe
m_overnight=pd.concat([m_overnight_a1_plus,m_overnight_a1,m_overnight_a2,m_overnight_a3],ignore_index=True,axis=0)
m_2_30= pd.concat([m_2_30_a1_plus,m_2_30_a1,m_2_30_a2 ,m_2_30_a3],ignore_index=True,axis=0)
m_31_60=pd.concat([m_31_60_a1_plus,m_31_60_a1,m_31_60_a2 ,m_31_60_a3],ignore_index=True,axis=0)
m_61_90=pd.concat([m_61_90_a1_plus,m_61_90_a1,m_61_90_a2 ,m_61_90_a3],ignore_index=True,axis=0)
m_91_120=pd.concat([m_91_120_a1_plus,m_91_120_a1,m_91_120_a2 ,m_91_120_a3],ignore_index=True,axis=0)
m_121_180=pd.concat([m_121_180_a1_plus,m_121_180_a1,m_121_180_a2 ,m_121_180_a3],ignore_index=True,axis=0)
m_181_270=pd.concat([m_181_270_a1_plus,m_181_270_a1,m_181_270_a2 ,m_181_270_a3],ignore_index=True,axis=0)
m_270_more=pd.concat([m_270_more_a1_plus,m_270_more_a1,m_270_more_a2 ,m_270_more_a3],ignore_index=True,axis=0)

### drop Maturity
def drop_ma(bond):
    bond=bond.drop(['Maturity'],axis=1)
    return bond


m_overnight=drop_ma(m_overnight)
m_2_30= drop_ma(m_2_30)
m_31_60=drop_ma(m_31_60)
m_61_90=drop_ma(m_61_90)
m_91_120=drop_ma(m_91_120)
m_121_180=drop_ma(m_121_180)
m_181_270=drop_ma(m_181_270)
m_270_more=drop_ma(m_270_more)

space= pd.DataFrame([np.nan*1000])
# space.rename(columns={'0':np.nan}, inplace=True)

##concat to one dataframe
m1=pd.concat([m_overnight,space,m_2_30,space,m_31_60,space,
              m_61_90,space,m_91_120,space,m_121_180,space,m_181_270,space,m_270_more],ignore_index=False,axis=1)

import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment,colors
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
# excel editor
wb = Workbook()
ws = wb.active
#merge_list=['A1:F1','H1:M1','O1:T1','V1:AA1','AC1:AH1','AJ1:AO1','AQ1:AV1','AX1:BC1']
#ws.title = "S&P ranking of CP"
#def mg(col):
    #for i in col:
        #ws.merge_cells(i)


#mg(merge_list)



#merge for the title
c1 = ws['A1']
c1.value = "Maturity: Overnight"

c2=ws['H1']
c2.value='Maturity: 2-30 days'

c3 =ws['O1']
c3.value='Maturity: 31-60 days'

c4=ws['V1']
c4.value = 'Maturity: 61-90 days'

c5=ws['AC1']
c5.value = 'Maturity: 91-120 days'

c6=ws['AJ1']
c6.value = 'Maturity: 121-180 days'

c7=ws['AQ1']
c7.value = 'Maturity: 181-270 days'

c8=ws['AX1']
c8.value = 'Maturity: More than 270 days'

##fill color and font color and alignment
mc =[c1,c2,c3,c4,c5,c6,c7,c8]
for e in mc:
    e.fill =PatternFill("solid", fgColor="2E86C1")
    e.font  = Font(b=True, color="FDFEFE")
    e.alignment = Alignment(horizontal="center", vertical="center")

#temp_list=['m_overnight','space','m_2_30','space','m_31_60','space','m_61_180','space','m_181_270','space','m_270_more']





# write main content into excel

for row in dataframe_to_rows(m1,index=False,header =True):
    ws.append(row)

# Format using a formula(fill blue)

#blue_fill=PatternFill(start_color='62bce5',
             #end_color='62bce5',
             #fill_type='solid')
#
#sp_list= ['A-1+','A-1','A-2','A-3']
#for i in sp_list:
       #ws.conditional_formatting.add('A1:A200',
           # FormulaRule(formula=[i], stopIfTrue=True, fill=blue_fill))


#a1 = ws['A3:A12']
#for i in a1:
    #a1[i].value
#ft = PatternFill(fill_type=None,
                 #start_color='62bce5',
                 #end_color='62bce5')
#a1.fill = ft

def set_color(bond):
    for i in bond:
        ws[i].fill = PatternFill("solid", fgColor="2E86C1")
        ws[i].font = Font(b=True, color="FDFEFE")


temp1= ['A3','A4','A5','A6','A7','A8','A9','A10','A11','A12',
        'A14','A15','A16','A17','A18','A19','A20','A21','A22','A23',
        'A25','A26','A27','A28','A29','A30','A31','A32','A33','A34',
        'A36','A37', 'A38','A39','A40','A41','A42','A43','A44','A45']

temp2= ['H3','H4','H5','H6','H7','H8','H9','H10','H11','H12',
        'H14','H15','H16','H17','H18','H19','H20','H21','H22','H23',
        'H25','H26','H27','H28','H29','H30','H31','H32','H33','H34',
        'H36','H37', 'H38','H39','H40','H41','H42','H43','H44','H45']

temp3= ['O3','O4','O5','O6','O7','O8','O9','O10','O11','O12',
        'O14','O15','O16','O17','O18','O19','O20','O21','O22','O23',
        'O25','O26','O27','O28','O29','O30','O31','O32','O33','O34',
        'O36','O37', 'O38','O39','O40','O41','O42','O43','O44','O45']


temp4= ['V3','V4','V5','V6','V7','V8','V9','V10','V11','V12',
        'V14', 'V15', 'V16', 'V17', 'V18', 'V19', 'V20', 'V21', 'V22', 'V23',
        'V25', 'V26', 'V27', 'V28', 'V29', 'V30', 'V31', 'V32', 'V33', 'V34',
        'V36', 'V37', 'V38', 'V39', 'V40', 'V41', 'V42', 'V43', 'V44','V45']

temp5 =['AC3','AC4','AC5','AC6','AC7','AC8','AC9','AC10','AC11','AC12',
        'AC14','AC15','AC16','AC17','AC18','AC19','AC20','AC21','AC22','AC23',
        'AC25','AC26','AC27','AC28','AC29','AC30','AC31','AC32','AC33','AC34',
        'AC36','AC37','AC38','AC39','AC40','AC41','AC42','AC43','AC44','AC45']

temp6 =['AJ3','AJ4','AJ5','AJ6','AJ7','AJ8','AJ9','AJ10','AJ11','AJ12',
        'AJ14', 'AJ15', 'AJ16', 'AJ17', 'AJ18', 'AJ19', 'AJ20', 'AJ21', 'AJ22', 'AJ23',
        'AJ25', 'AJ26', 'AJ27', 'AJ28', 'AJ29', 'AJ30', 'AJ31', 'AJ32', 'AJ33', 'AJ34',
        'AJ36', 'AJ37', 'AJ38', 'AJ39', 'AJ40', 'AJ41', 'AJ42', 'AJ43', 'AJ44','AJ45']

temp7 =['AQ3','AQ4','AQ5','AQ6','AQ7','AQ8','AQ9','AQ10','AQ11','AQ12',
        'AQ14', 'AQ15', 'AQ16', 'AQ17', 'AQ18', 'AQ19', 'AQ20', 'AQ21', 'AQ22', 'AQ23',
        'AQ25', 'AQ26', 'AQ27', 'AQ28', 'AQ29', 'AQ30', 'AQ31', 'AQ32', 'AQ33', 'AQ34',
        'AQ36', 'AQ37', 'AQ38', 'AQ39', 'AQ40', 'AQ41', 'AQ42', 'AQ43', 'AQ44','AQ45']

temp8 =['AX3','AX4','AX5','AX6','AX7','AX8','AX9','AX10','AX11','AX12',
        'AX14', 'AX15', 'AX16', 'AX17', 'AX18', 'AX19', 'AX20', 'AX21', 'AX22', 'AX23',
        'AX25', 'AX26', 'AX27', 'AX28', 'AX29', 'AX30', 'AX31', 'AX32', 'AX33', 'AX34',
        'AX36', 'AX37', 'AX38', 'AX39', 'AX40', 'AX41', 'AX42', 'AX43', 'AX44','AX45']


set_color(temp1)
set_color(temp2)
set_color(temp3)
set_color(temp4)
set_color(temp5)
set_color(temp6)
set_color(temp7)
set_color(temp8)

def adjust(bond):
    for i in bond:
        ws[i].value=np.nan

temp=['G2','N2','U2','AB2','AI2','AP2','AW2']

adjust(temp)



for i in range(7):
        ws.insert_rows(1)


# create an image
img = Image('logo.png')
 # add to worksheet and anchor next to cells
ws.add_image(img, 'A1')

#move the buckets
ws.move_range("V8:AO52", rows=46, cols=-21)

ws.move_range("AQ8:BC52", rows=92, cols=-42)



##alignment
##a_list= ['D9',
   ## ,'K','R']

##for i in a_list:
   ## ws[i].alignment =Alignment(horizontal="center")

#####insert lines
for i in range(2):
        ws.insert_rows(21)
for i in range(2):
        ws.insert_rows(34)

for i in range(2):
        ws.insert_rows(47)

for i in range(2):
        ws.insert_rows(73)

for i in range(2):
        ws.insert_rows(86)

for i in range(2):
        ws.insert_rows(99)

for i in range(2):
        ws.insert_rows(125)

for i in range(2):
        ws.insert_rows(138)

for i in range(2):
        ws.insert_rows(151)

# copy maturity
col_num1=[1,8,15]
row_num1=[21,34,47]

col_num2=[1,8,15]
row_num2=[73,86,99]

col_num3=[1,8]
row_num3=[125,138,151]

def copy_m(col,row):
    for i in col:
        for j in row:
            ws.cell(column=i,row =j).value=ws.cell(column=i,row=row[0]-13).value


copy_m(col_num1,row_num1)
copy_m(col_num2,row_num2)
copy_m(col_num3,row_num3)



#merge title
merge_list=['A8:F8','H8:M8','O8:T8',
            'A60:F60','H60:M60','O60:T60',
            'A112:F112','H112:M112','O112:T112']


ws.title = "CP-Top10 Yield by MTY"
def mg(col):
    for i in col:
        ws.merge_cells(i)




mg(merge_list)




###additional decoration of cp top 10

#for the maturity 0 to 60 section
# 9~19  ,23~32, 36~45, 49~58
temp =['D9','D10','D11','D12','D13','D14','D15','D16','D17','D18','D19']

def alg(col):
    for i in col:

      ws[i].alignment = Alignment(shrink_to_fit=True)




#for the maturity 61 to 180 section
# 61~90  ,23~32, 36~45, 49~58








output_file_name= input('Please provide the filename you wanna save as(strict format: name.xlsx): ')


wb.save(output_file_name)
